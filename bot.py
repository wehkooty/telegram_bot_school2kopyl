import logging
import asyncio
from datetime import datetime
import os
import openpyxl

from aiogram import Bot, Dispatcher, types, F
from aiogram.filters import CommandStart
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
from aiogram.enums import ParseMode
from aiogram.client.default import DefaultBotProperties

import config

# --- Логирование ---
logging.basicConfig(level=logging.INFO)

# --- Бот ---
bot = Bot(
    token=config.API_TOKEN,
    default=DefaultBotProperties(parse_mode=ParseMode.HTML)
)
dp = Dispatcher()

# --- Хранилище ---
pending_answers = {}

# --- Excel ---
def init_excel():
    filename = "requests.xlsx"
    if not os.path.exists(filename):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(
            ["ID", "Дата/время", "Имя", "Username", "Тип", "Вопрос", "Ответ администратора"]
        )
        wb.save(filename)


def save_question(user, text, msg_type):
    init_excel()
    filename = "requests.xlsx"
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    new_id = ws.max_row
    ws.append(
        [
            new_id,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            user.full_name,
            f"@{user.username}" if user.username else "—",
            msg_type,
            text,
            "",
        ]
    )
    wb.save(filename)
    return new_id


def save_answer(row_index, answer_text):
    filename = "requests.xlsx"
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    ws.cell(row=row_index, column=7).value = answer_text
    wb.save(filename)


# --- Главное меню ---
def parent_menu():
    """Меню для родителей"""
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="❓ Задать вопрос", callback_data="ask_question")],
            [InlineKeyboardButton(text="📅 Записаться на услугу", callback_data="service")],
            [InlineKeyboardButton(text="📞 Контакты", callback_data="contacts")],
            [InlineKeyboardButton(text="📜 Мои вопросы", callback_data="my_questions")],
        ]
    )


def admin_menu():
    """Меню для администрации"""
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="📋 Список вопросов", callback_data="list_questions")],
            [InlineKeyboardButton(text="📊 Статистика", callback_data="stats")],
            [InlineKeyboardButton(text="📂 Экспорт в Excel", callback_data="export")],
        ]
    )


# --- /start ---
@dp.message(CommandStart())
async def start(message: types.Message):
    if message.from_user.id == config.ADMIN_ID:
        await message.answer(
            "👋 Добро пожаловать, <b>администратор</b>!\nВыберите действие:",
            reply_markup=admin_menu(),
        )
    else:
        await message.answer(
            "👋 Здравствуйте! Я школьный помощник для родителей 🏫\nВыберите действие:",
            reply_markup=parent_menu(),
        )


# --- Обработка меню родителя ---
@dp.callback_query(F.data == "ask_question")
async def cb_ask_question(callback: types.CallbackQuery):
    await callback.message.answer("✍ Напишите свой вопрос, и мы передадим его администрации школы.")
    await callback.answer()


@dp.callback_query(F.data == "service")
async def cb_service(callback: types.CallbackQuery):
    await callback.message.answer("📅 Укажите, на какую услугу хотите записаться (кружок, консультация).")
    await callback.answer()


@dp.callback_query(F.data == "contacts")
async def cb_contacts(callback: types.CallbackQuery):
    await callback.message.answer("📞 Телефон школы: 80171955108\n📧 Email: sh2@kopyl.gov.by")
    await callback.answer()


@dp.callback_query(F.data == "my_questions")
async def cb_my_questions(callback: types.CallbackQuery):
    init_excel()
    wb = openpyxl.load_workbook("requests.xlsx")
    ws = wb.active
    user = callback.from_user.full_name

    result = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        qid, date, name, username, msg_type, question, answer = row
        if user == name:
            result.append(
                f"№{qid} ({msg_type}, {date}):\n❓ {question}\n➡️ {answer if answer else '⏳ В ожидании'}"
            )

    if not result:
        await callback.message.answer("📭 У вас пока нет вопросов.")
    else:
        await callback.message.answer("\n\n".join(result))
    await callback.answer()


# --- Обработка меню админа ---
@dp.callback_query(F.data == "list_questions")
async def cb_list_questions(callback: types.CallbackQuery):
    if callback.from_user.id != config.ADMIN_ID:
        await callback.answer("⛔ Доступ запрещён", show_alert=True)
        return

    init_excel()
    wb = openpyxl.load_workbook("requests.xlsx")
    ws = wb.active

    text = "📋 <b>Список вопросов</b>\n\n"
    kb = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        qid, date, name, username, msg_type, question, answer = row
        if msg_type == "Вопрос" and not answer:
            kb.append([InlineKeyboardButton(text=f"✍ Ответить на №{qid}", callback_data=f"reply_{row[0]}_{qid}")])
            text += f"№{qid} от {name} ({username})\n❓ {question}\n\n"

    if not kb:
        await callback.message.answer("✅ Нет новых вопросов.")
    else:
        await callback.message.answer(text, reply_markup=InlineKeyboardMarkup(inline_keyboard=kb))
    await callback.answer()


@dp.callback_query(F.data == "stats")
async def cb_stats(callback: types.CallbackQuery):
    if callback.from_user.id != config.ADMIN_ID:
        await callback.answer("⛔ Доступ запрещён", show_alert=True)
        return

    init_excel()
    wb = openpyxl.load_workbook("requests.xlsx")
    ws = wb.active

    total, questions, services, closed = 0, 0, 0, 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        total += 1
        msg_type = row[4]
        answer = row[6]
        if msg_type == "Вопрос":
            questions += 1
            if answer:
                closed += 1
        if msg_type == "Заявка":
            services += 1

    await callback.message.answer(
        f"📊 <b>Статистика</b>\n\n"
        f"Всего сообщений: {total}\n"
        f"❓ Вопросов: {questions} (✅ {closed} закрыто)\n"
        f"📅 Заявок: {services}\n"
    )
    await callback.answer()


@dp.callback_query(F.data == "export")
async def cb_export(callback: types.CallbackQuery):
    if callback.from_user.id != config.ADMIN_ID:
        await callback.answer("⛔ Доступ запрещён", show_alert=True)
        return

    if os.path.exists("requests.xlsx"):
        await callback.message.answer_document(open("requests.xlsx", "rb"))
    else:
        await callback.message.answer("⛔ Файл пока не создан.")
    await callback.answer()


# --- Обработка ответов админа ---
@dp.callback_query(F.data.startswith("reply_"))
async def cb_reply(callback: types.CallbackQuery):
    if callback.from_user.id == config.ADMIN_ID:
        _, parent_id, row_index = callback.data.split("_")
        parent_id, row_index = int(parent_id), int(row_index)

        pending_answers[config.ADMIN_ID] = (parent_id, row_index)

        # Сообщение админу
        await callback.message.answer(f"✍ Введите ответ для вопроса №{row_index}.")

        # Сообщение родителю
        try:
            await bot.send_message(
                parent_id,
                f"📢 Ваш вопрос №{row_index} принят администрацией в обработку. "
                "Ожидайте ответа ⏳"
            )
        except:
            await callback.message.answer("⚠️ Не удалось уведомить родителя.")

        await callback.answer()
    else:
        await callback.answer("⛔ Доступ запрещён", show_alert=True)


# --- Сообщения пользователей ---
@dp.message()
async def handler(message: types.Message):
    user = message.from_user
    text = message.text

    if user.id == config.ADMIN_ID and user.id in pending_answers:
        parent_id, row_index = pending_answers.pop(user.id)
        save_answer(row_index, text)
        await bot.send_message(parent_id, f"📩 Ответ от администрации:\n{text}")
        await message.answer("✅ Ответ отправлен родителю.")
        return

    if user.id != config.ADMIN_ID:
        if "запис" in text.lower() or "услуг" in text.lower():
            msg_type = "Заявка"
        elif "?" in text:
            msg_type = "Вопрос"
        else:
            msg_type = "Другое"

        qid = save_question(user, text, msg_type)

        if msg_type == "Вопрос":
            kb = InlineKeyboardMarkup(
                inline_keyboard=[
                    [InlineKeyboardButton(text=f"✍ Ответить (№{qid})", callback_data=f"reply_{user.id}_{qid}")]
                ]
            )
            await bot.send_message(
                config.ADMIN_ID,
                f"❓ Вопрос №{qid} от {user.full_name} (@{user.username}):\n{text}",
                reply_markup=kb,
            )
            await message.answer(f"✅ Ваш вопрос зарегистрирован (№{qid}). Ожидайте ответа.")
        elif msg_type == "Заявка":
            await bot.send_message(
                config.ADMIN_ID,
                f"📅 Заявка №{qid} от {user.full_name} (@{user.username}):\n{text}",
            )
            await message.answer(f"✅ Ваша заявка зарегистрирована (№{qid}). Администрация свяжется с вами.")
        else:
            await bot.send_message(
                config.ADMIN_ID,
                f"💬 Сообщение №{qid} от {user.full_name} (@{user.username}):\n{text}",
            )
            await message.answer("✅ Сообщение доставлено администрации школы.")


# --- Запуск ---
async def main():
    init_excel()
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())
